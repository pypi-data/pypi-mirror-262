import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
import threading
import web3
import time
import requests
import openpyxl
import re


def urljoin(*args):
    return "\\".join(map(lambda x: str(x).rstrip('\\'), args))


def regexp(exp, mes, arg=''):
    while not re.match(exp, arg):
        arg = input(mes)
    return arg


def urlcheck(surl):
    r = requests.get(surl)
    return r.status_code


def excel_open(s_filepath):
    book = openpyxl.load_workbook(s_filepath + '.xlsx')
    return book


def excel_sheet(book):
    sheet = book['sides']
    return sheet


def new_tabs_ch(driver):
    b = driver.current_window_handle
    for handle in driver.window_handles:
        if handle != b:
            driver.switch_to.window(handle)
            driver.close()
            driver.switch_to.window(b)

def meta_reg(driver, sheet, spass, rn, delay, fl):
    try:
        element = WebDriverWait(driver, delay).\
            until(ec.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/a')))
    except TimeoutException:
        print('Hello page not loaded...')
    else:
        element.click()
    if fl == "y":
        try:
            element = WebDriverWait(driver, delay).\
                until(ec.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div/div[1]/a')))
        except TimeoutException:
            print('Hello page not loaded...')
        else:
            element.click()
    else:
        try:
            element = WebDriverWait(driver, delay).\
                until(ec.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div/div/div[2]/a')))
        except TimeoutException:
            print('Hello page not loaded...')
        else:
            element.click()


def run_chrome(b_add_wallets):
    root_dir = os.path.dirname(os.path.abspath(__file__))
    metamask_path = urljoin(root_dir, r'MetaMask\Sui Wallet  0.2.1.0.crx')
    excel_path = urljoin(root_dir, r'Sid_Phrases\sides')
    ch_dr_path = urljoin(root_dir, r'chromedriver_win32\chromedriver.exe')
    profiles_path = urljoin(root_dir, r'Data_profiles\profile_')
    metamask_extension_url = 'chrome-extension://opcgpfmipidbgpenhmajoajpbobppdil/ui.html#/welcome'

    book = excel_open(excel_path)
    sheet = excel_sheet(book)
    #print(str(sheet.max_row))
    #for row in range(1,sheet.max_row + 1):
    #    print(sheet.cell(row, 1).value)
    rn = sheet.max_row
    #input()
    qty_wallets = 0

    book = excel_open(excel_path)
    sheet = excel_sheet(book)
    rn = sheet.max_row
    options = Options()
    options.add_argument('start-maximized')
    options.add_argument('user-data-dir='+profiles_path)
    options.add_extension(metamask_path)


    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    delay = 10  # seconds
    driver.get(metamask_extension_url)
    time.sleep(10)
    new_tabs_ch(driver)
    meta_reg(driver, 1, '1', 1, delay, b_add_wallets)
    #flurl = urlcheck('https://google.com')
    select_network = "bsc"
    #spass = s_pass(sheet, rn, 0)
    #if driver.current_url == metamask_extension_url:
    #    meta_reg(driver, sheet, spass, rn, delay)
    #    for s_network in ['bsc', 'polygon', 'cronos', 'optimism', 'arbitrum', 'heco', 'moonriver', 'okex', 'aurora', 'avalanche', 'boba']:
    #        meta_network_reg(driver, delay, s_network)
    #else:
    #    meta_unl(driver, spass, delay)
    #meta_network_select(driver, delay, select_network)
    input()
    #driver.close()


qty_wallets = int(regexp(r'^\d+$', 'Write number wallets for work: '))
b_add_wallets = regexp(r'^[yn]$', 'Do You want create New wallets?(y/n) ')
threads = []

for i in range(qty_wallets):
    t = threading.Thread(target=run_chrome, args=(b_add_wallets,))
    t.start()
    threads.append(t)

for t in threads:
    t.join()
